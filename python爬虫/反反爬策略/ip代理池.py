import random

import requests
from lxml import etree
import openpyxl
from fake_useragent import UserAgent

# 实例化 user-agent 对象
ua = UserAgent(verify_ssl=False)
# headers = {
#     "User-Agent": ua.random}
# 不同的代理IP,代理ip的类型必须和请求url的协议头保持一致
proxy_list = [
     {"http": "112.115.57.20:3128"},
     {'http': '121.41.171.223:3128'}
]

# 随机获取代理IP
proxy = random.choice(proxy_list)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.79 Safari/537.36',
}
excelFilePath = "ip_File.xlsx"
sheetName = "sheet1"


# 获取ip
def getinfo(url):
    html = requests.get(url=url, headers=headers,proxies=proxy, timeout=30)
    # print(html.status_code)
    root = etree.HTML(html.content)
    ip_list = []
    # 测试：仅爬取每页前五ip，真正爬取时可放宽
    for tr in root.xpath("//*[@id='list']/table/tbody/tr"):
        ip_address = tr.xpath("td[1]/text()")[0]  # ip地址
        ip_port = tr.xpath("td[2]/text()")[0]  # ip端口号
        ip_protocol = tr.xpath("td[4]/text()")[0]  # ip网络协议
        ip = "{}://{}:{}".format(ip_protocol, ip_address, ip_port)
        ip_s="{}:{}".format(ip_address, ip_port)

        proxies_dict={}
        # 构建代理ip字典
        proxies_dict[ip_protocol] = ip_address + ':' + ip_port
        # print(proxies_dict)
        # 有用则收录起来
        if verify_ip(proxies_dict):
            ip_list.append([ip_protocol, ip_s])
            print('true')
        else:
            print('flase')

    write_excel(excelFilePath=excelFilePath, values=ip_list)  # 写入数据


# 验证ip有效
def verify_ip(proxies):
    try:
        response = requests.get('http://www.baidu.com', headers=headers, proxies=proxies, timeout=0.1)  # 超时报错
        if response.status_code == 200:
            print(proxies, "            OK")
            return True
        else:
            return False
    except Exception as error:
        print(error)



# 追加填写数据
def write_excel(excelFilePath, values):
    data = openpyxl.load_workbook(excelFilePath)
    table = data.active
    nrows = table.max_row  # 获得行数
    # 注意 行 列下标是从1开始的
    for i in range(1, len(values) + 1):  # 行
        for j in range(1, len(values[i - 1]) + 1):  # 列
            table.cell(nrows + i, j).value = values[i - 1][j - 1]
    data.save(excelFilePath)


# 写入excel表头
def write_excel_title(excelFilePath, sheetName, datas):
    index = len(datas)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheetName
    for i in range(0, index):
        for j in range(0, len(datas[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(datas[i][j]))
    workbook.save(excelFilePath)
    print("表头写入成功！")


# 读出excel
def read_excel(excelFilePath, sheetName):
    workbook = openpyxl.load_workbook(excelFilePath)
    sheet = workbook[sheetName]
    # print(sheet.rows[0].value)
    ip_list = []
    for row in sheet.rows:
        ip_list.append([row[0].value, row[1].value])
    return ip_list[1:]  # 剔除表头


# 翻页
def nextPage(url_model):
    # 获取前10页的代理ip
    for i in range(1, 2):
        url = url_model.format(i)
        getinfo(url)


if __name__ == '__main__':
    # url模式化
    # 快代理
    url_model_list = ["https://www.kuaidaili.com/free/inha/{}",  # 国内高匿代理
                      "https://www.kuaidaili.com/free/intr/{}"  # 国内普通代理
                      ]

    # 写入表格标题 重置表格
    write_excel_title(excelFilePath=excelFilePath, sheetName=sheetName, datas=[['Http/https', 'ip_address']])

    '''
    多线程实现，四个网站同步爬取，尚未完成
    '''
    nextPage(url_model_list[0])
    nextPage(url_model_list[1])
    # nextPage(url_model_list[2])
    # nextPage(url_model_list[3])

    # urla="https://www.kuaidaili.com/free/intr/{}"
    # nextPage(urla)

    # 代理ip的使用
    ip_list = read_excel(excelFilePath=excelFilePath, sheetName=sheetName)
    for ip in ip_list:
        print({ip[0]:ip[1]})







